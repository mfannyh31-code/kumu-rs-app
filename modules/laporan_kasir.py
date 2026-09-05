import streamlit as st
import pandas as pd
import io
import streamlit.components.v1 as components
from datetime import datetime
from db import get_db, format_rupiah
from sqlalchemy import text

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def render_page():
    if 'show_report_kasir' not in st.session_state:
        st.session_state.show_report_kasir = False
    if 'active_date' not in st.session_state:
        st.session_state.active_date = datetime.today().date()
    if 'active_shift' not in st.session_state:
        st.session_state.active_shift = "Pagi"
    
    if 'active_kasir' not in st.session_state:
        if st.session_state.role in ["Bendahara", "Super Admin"]:
            st.session_state.active_kasir = "Semua Kasir"
        else:
            st.session_state.active_kasir = str(st.session_state.user).upper()

    # --- CSS KHUSUS UNTUK LAPORAN (DENGAN KELOMPOK WARNA PIUTANG) ---
    st.markdown("""
        <style>
        .report-header {
            font-size: 24px;
            font-weight: 800;
            color: #0F172A;
            margin-bottom: 0px;
        }
        .report-subheader {
            font-size: 13px;
            color: #64748B;
            margin-bottom: 20px;
            font-weight: 600;
        }
        .report-card {
            background-color: #FFFFFF;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.02);
            border: 1px solid #E2E8F0;
        }
        .btn-filter button {
            background-color: #0084FF !important;
            color: white !important;
            border-radius: 6px !important;
            height: 42px !important;
            margin-top: 28px !important;
            font-weight: bold !important;
        }
        
        .btn-pdf button, div[data-testid="stDownloadButton"] > button { 
            background-color: #009688 !important; 
            color: white !important; 
            font-weight: bold !important; 
            border-radius: 6px !important;
            width: 100% !important;
            border: none !important;
            padding: 8px 14px !important;
            height: auto !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1) !important;
        }
        
        .tbl-report {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            margin-bottom: 20px;
            font-size: 13.5px;
            color: #334155;
            border: 1px solid #CBD5E1;
        }
        .tbl-report th {
            background-color: #F8FAFC;
            border: 1px solid #CBD5E1;
            padding: 10px 8px;
            color: #0F172A;
            font-weight: 700;
            text-align: left;
        }
        .tbl-report td {
            padding: 10px 8px;
            border: 1px solid #CBD5E1;
        }
        .tbl-report tr.total-row td {
            font-weight: 800;
            color: #0F172A;
            background-color: #F1F5F9;
            border: 1px solid #CBD5E1;
        }
        
        .header-transaksi {
            background-color: #48BB78;
            color: white;
            text-align: center;
            font-weight: bold;
            padding: 10px;
            font-size: 14px;
            letter-spacing: 1px;
            border-top-left-radius: 6px;
            border-top-right-radius: 6px;
            margin-top: 20px;
        }
        
        .transaksi-meta {
            background-color: #E6FFFA;
            color: #234E52;
            padding: 8px 12px;
            font-weight: 700;
            font-size: 13.5px;
            border-left: 4px solid #319795;
            border-top: 1px solid #CBD5E1;
            border-right: 1px solid #CBD5E1;
            margin-bottom: 0px;
        }
        
        .summary-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
            font-weight: 700;
            color: #334155;
            margin-top: 10px;
            border: 1px solid #CBD5E1;
        }
        .summary-table td {
            padding: 8px 12px;
            text-align: right;
            border: 1px solid #CBD5E1;
        }
        .summary-table td:first-child {
            width: 80%;
            color: #0F172A;
            text-align: left;
        }
        .piutang-row td {
            background-color: #EFF6FF !important;
            color: #1E40AF !important;
        }
        .summary-table tr.setoran-row td {
            background-color: #E2E8F0;
            font-size: 15px;
            font-weight: 800;
            color: #0F172A;
        }
        
        @media print {
            .btn-pdf, .btn-excel, .btn-filter, section[data-testid="stSidebar"], header {
                display: none !important;
            }
            .report-card { border: none !important; box-shadow: none !important; padding: 0 !important; }
        }
        </style>
    """, unsafe_allow_html=True)

    with get_db() as conn:
        try:
            df_users = pd.read_sql_query(text("SELECT DISTINCT cashier_username FROM transactions WHERE cashier_username IS NOT NULL"), conn)
            list_users = ["Semua Kasir"] + [str(u).upper() for u in df_users['cashier_username'].tolist() if str(u).strip() != ""]
        except:
            list_users = ["Semua Kasir"]

    # --- HEADER ---
    st.markdown('<div class="report-header">Laporan Penerimaan Kasir</div>', unsafe_allow_html=True)
    
    if st.session_state.show_report_kasir:
        st.markdown(f'<div class="report-subheader">Tanggal: {st.session_state.active_date.strftime("%d-%b-%Y")} | Shift: {st.session_state.active_shift} | Kasir: {st.session_state.active_kasir}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="report-subheader">Silakan atur filter dan klik tombol Filter Data</div>', unsafe_allow_html=True)

    st.markdown('<div class="report-card">', unsafe_allow_html=True)

    # --- FILTER SECTION ---
    f1, f2, f3, f4 = st.columns([1.5, 1.5, 1.5, 3])
    with f1:
        filter_date = st.date_input("Tgl. Transaksi*", value=st.session_state.active_date)
    with f2:
        shift_options = ["Pagi", "Sore", "Malam", "Semua Shift"]
        idx_shift = shift_options.index(st.session_state.active_shift) if st.session_state.active_shift in shift_options else 0
        filter_shift = st.selectbox("Shift*", shift_options, index=idx_shift)
    with f3:
        if st.session_state.role in ["Bendahara", "Super Admin"]:
            idx_ksr = list_users.index(st.session_state.active_kasir) if st.session_state.active_kasir in list_users else 0
            filter_user = st.selectbox("Kasir*", list_users, index=idx_ksr)
        else:
            filter_user = str(st.session_state.user).upper()
            st.text_input("Kasir*", value=filter_user, disabled=True)
            
    with f4:
        st.markdown('<div class="btn-filter">', unsafe_allow_html=True)
        if st.button("Filter Data", use_container_width=True):
            st.session_state.show_report_kasir = True
            st.session_state.active_date = filter_date
            st.session_state.active_shift = filter_shift
            st.session_state.active_kasir = filter_user
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # =====================================================================
    # RENDER LAPORAN UTAMA
    # =====================================================================
    if st.session_state.show_report_kasir:
        current_date = st.session_state.active_date
        current_shift = st.session_state.active_shift
        current_kasir = st.session_state.active_kasir
        str_date = current_date.strftime("%d-%b-%Y")
        
        d1 = f"{current_date}%"                                    
        d2 = current_date.strftime('%d/%m/%Y') + "%"              
        d3 = current_date.strftime('%d/%b/%Y') + "%"              
        d4 = current_date.strftime('%d-%b-%Y') + "%"              
        
        with get_db() as conn:
            # 1. Query Transaksi Item dengan Integrasi Hirarki Master Data
            query_items = """
                SELECT i.category_name, i.action_name, i.qty, i.subtotal, t.receipt_date, t.cashier_username, t.shift,
                       COALESCE(sc.name, 'Lainnya') as service_category_name
                FROM transaction_items i
                JOIN transactions t ON i.receipt_no = t.receipt_no
                LEFT JOIN categories c ON i.category_name = c.name
                LEFT JOIN service_categories sc ON c.service_category_id = sc.id
                WHERE (t.receipt_date LIKE :d1 OR t.receipt_date LIKE :d2 OR t.receipt_date LIKE :d3 OR t.receipt_date LIKE :d4)
            """
            params_items = {"d1": d1, "d2": d2, "d3": d3, "d4": d4}
            if current_shift != "Semua Shift":
                query_items += " AND UPPER(TRIM(t.shift)) = UPPER(TRIM(:shf))"
                params_items["shf"] = current_shift
            if current_kasir != "Semua Kasir":
                query_items += " AND UPPER(TRIM(t.cashier_username)) = UPPER(TRIM(:ksr))"
                params_items["ksr"] = current_kasir
            df_items = pd.read_sql_query(text(query_items), conn, params=params_items)

            # 2. Query Transaksi Induk
            query_tx = """
                SELECT payment_method, final_amount, pay_tunai, pay_transfer, pay_edc, pay_qris, pay_va, pay_pengakuan_bendahara 
                FROM transactions 
                WHERE (receipt_date LIKE :d1 OR receipt_date LIKE :d2 OR receipt_date LIKE :d3 OR receipt_date LIKE :d4)
            """
            params_tx = {"d1": d1, "d2": d2, "d3": d3, "d4": d4}
            if current_shift != "Semua Shift":
                query_tx += " AND UPPER(TRIM(shift)) = UPPER(TRIM(:shf))"
                params_tx["shf"] = current_shift
            if current_kasir != "Semua Kasir":
                query_tx += " AND UPPER(TRIM(cashier_username)) = UPPER(TRIM(:ksr))"
                params_tx["ksr"] = current_kasir
            df_tx = pd.read_sql_query(text(query_tx), conn, params=params_tx)

            # 3. Query Deposit / Uang Muka
            query_depo = "SELECT amount, payment_method FROM deposits WHERE (deposit_date LIKE :d1 OR deposit_date LIKE :d2 OR deposit_date LIKE :d3 OR deposit_date LIKE :d4)"
            params_depo = {"d1": d1, "d2": d2, "d3": d3, "d4": d4}
            if current_shift != "Semua Shift":
                query_depo += " AND UPPER(TRIM(shift)) = UPPER(TRIM(:shf))"
                params_depo["shf"] = current_shift
            if current_kasir != "Semua Kasir":
                query_depo += " AND UPPER(TRIM(input_by)) = UPPER(TRIM(:ksr))"
                params_depo["ksr"] = current_kasir
            df_depo = pd.read_sql_query(text(query_depo), conn, params=params_depo)

            # 4. Query Pembayaran Piutang
            query_piu = """
                SELECT r.patient_name as debtor_name, p.amount, p.method, p.shift, p.input_by, p.pay_date 
                FROM receivables_payments p
                LEFT JOIN receivables r ON p.debt_id = r.id
                WHERE (p.pay_date LIKE :d1 OR p.pay_date LIKE :d2 OR p.pay_date LIKE :d3 OR p.pay_date LIKE :d4)
            """
            params_piu = {"d1": d1, "d2": d2, "d3": d3, "d4": d4}
            if current_shift != "Semua Shift":
                query_piu += " AND UPPER(TRIM(p.shift)) = UPPER(TRIM(:shf))"
                params_piu["shf"] = current_shift
            if current_kasir != "Semua Kasir":
                query_piu += " AND UPPER(TRIM(p.input_by)) = UPPER(TRIM(:ksr))"
                params_piu["ksr"] = current_kasir
            try:
                df_piu = pd.read_sql_query(text(query_piu), conn, params=params_piu)
            except:
                df_piu = pd.DataFrame()

        def is_tunai(method_str):
            if not method_str: return False
            return "TUNAI" in str(method_str).upper().strip()

        piu_tunai = df_piu[df_piu['method'].apply(is_tunai)]['amount'].sum() if not df_piu.empty and 'amount' in df_piu.columns else 0.0
        piu_transfer = df_piu[df_piu['method'].str.upper().str.contains('TRANSFER', na=False)]['amount'].sum() if not df_piu.empty and 'amount' in df_piu.columns else 0.0
        piu_edc = df_piu[df_piu['method'].str.upper().str.contains('EDC', na=False)]['amount'].sum() if not df_piu.empty and 'amount' in df_piu.columns else 0.0
        piu_qris = df_piu[df_piu['method'].str.upper().str.contains('QRIS', na=False)]['amount'].sum() if not df_piu.empty and 'amount' in df_piu.columns else 0.0
        piu_va = df_piu[df_piu['method'].str.upper().str.contains('VIRTUAL', na=False) | df_piu['method'].str.upper().str.contains('VA', na=False)]['amount'].sum() if not df_piu.empty and 'amount' in df_piu.columns else 0.0

        tx_tunai = df_tx['pay_tunai'].sum() if not df_tx.empty and 'pay_tunai' in df_tx.columns else 0.0
        tx_edc = df_tx['pay_edc'].sum() if not df_tx.empty and 'pay_edc' in df_tx.columns else 0.0
        tx_qris = df_tx['pay_qris'].sum() if not df_tx.empty and 'pay_qris' in df_tx.columns else 0.0
        tx_va = df_tx['pay_va'].sum() if not df_tx.empty and 'pay_va' in df_tx.columns else 0.0
        tx_transfer = df_tx['pay_transfer'].sum() if not df_tx.empty and 'pay_transfer' in df_tx.columns else 0.0

        df_depo_masuk = df_depo[df_depo['amount'] > 0] if not df_depo.empty else pd.DataFrame()
        depo_tunai = df_depo_masuk[df_depo_masuk['payment_method'].apply(is_tunai)]['amount'].sum() if not df_depo_masuk.empty else 0.0
        depo_edc = df_depo_masuk[df_depo_masuk['payment_method'].str.upper().str.contains('EDC', na=False)]['amount'].sum() if not df_depo_masuk.empty else 0.0
        depo_transfer = df_depo_masuk[~df_depo_masuk['payment_method'].apply(is_tunai) & ~df_depo_masuk['payment_method'].str.upper().str.contains('EDC', na=False)]['amount'].sum() if not df_depo_masuk.empty else 0.0
        
        df_depo_keluar = df_depo[df_depo['amount'] < 0] if not df_depo.empty else pd.DataFrame()
        pengembalian_depo_tunai = df_depo_keluar[df_depo_keluar['payment_method'].apply(is_tunai)]['amount'].abs().sum() if not df_depo_keluar.empty else 0.0
        pengembalian_depo_transfer = df_depo_keluar[~df_depo_keluar['payment_method'].apply(is_tunai)]['amount'].abs().sum() if not df_depo_keluar.empty else 0.0

        tot_pengakuan = df_tx['pay_pengakuan_bendahara'].sum() if not df_tx.empty and 'pay_pengakuan_bendahara' in df_tx.columns else 0.0

        # =====================================================================
        # TAMPILAN: Uraian Pemasukan Per Unit Layanan
        # =====================================================================
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"#### Uraian Pemasukan Per Unit Layanan ({str_date})")
        
        html_unit = "<table class='tbl-report'>"
        html_unit += "<tr><th style='width: 10%;'>No</th><th style='width: 35%;'>Kategori Layanan Utama</th><th style='width: 35%;'>Unit Layanan</th><th style='width: 20%; text-align: right;'>Penerimaan<br>Jumlah</th></tr>"
        
        total_pendapatan = 0
        data_ex_unit = []
        if not df_items.empty:
            df_unit = df_items.groupby(['service_category_name', 'category_name'])['subtotal'].sum().reset_index()
            for idx, row in df_unit.iterrows():
                amt = float(row['subtotal'])
                total_pendapatan += amt
                html_unit += f"<tr><td>{idx + 1}</td><td style='font-weight:600;'>{row['service_category_name']}</td><td style='color:#6366F1; font-weight:600;'>{row['category_name']}</td><td style='text-align: right;'>{format_rupiah(amt).replace('Rp ', '')}</td></tr>"
                data_ex_unit.append({"No": idx + 1, "Kategori Layanan Utama": row['service_category_name'], "Unit Layanan": row['category_name'], "Pemasukan (Rp)": amt})
        else:
            html_unit += "<tr><td colspan='4' style='text-align:center;'>Tidak ada data pada filter ini</td></tr>"
            
        html_unit += f"<tr class='total-row'><td colspan='3'>Total</td><td style='text-align: right;'>{format_rupiah(total_pendapatan).replace('Rp ', '')}</td></tr>"
        html_unit += "</table>"
        st.markdown(html_unit, unsafe_allow_html=True)
        df_ex_unit = pd.DataFrame(data_ex_unit)

        # =====================================================================
        # TAMPILAN: Uraian Semua Data (Berdasarkan Jenis Tindakan)
        # =====================================================================
        st.markdown(f"#### Uraian Semua Data ({str_date})")
        
        display_user_meta = current_kasir if current_kasir != "Semua Kasir" else "SEMUA KASIR"
        display_shift_meta = current_shift if current_shift != "Semua Shift" else "SEMUA SHIFT"
        
        st.markdown(f"""
            <div class="transaksi-meta">
                👤 NAMA USER: <strong>{display_user_meta}</strong> &nbsp;|&nbsp; 🕒 SHIFT: <strong>{display_shift_meta}</strong>
            </div>
        """, unsafe_allow_html=True)
        
        html_semua = "<div class='header-transaksi'>TRANSAKSI</div>"
        html_semua += "<table class='tbl-report' style='margin-top:0px; border-top:none;'>"
        html_semua += "<tr><th style='width: 5%;'>No</th><th style='width: 15%;'>Tanggal Kuitansi</th><th style='width: 25%;'>Unit Layanan</th><th style='width: 35%;'>Uraian Jenis Tindakan</th><th style='width: 8%; text-align: center;'>Item</th><th style='width: 12%; text-align: right;'>Penerimaan<br>Jumlah</th></tr>"
        
        total_qty = 0
        total_jumlah_semua = 0
        data_ex_semua = []
        
        if not df_items.empty:
            df_action = df_items.groupby(['receipt_date', 'category_name', 'action_name']).agg({'qty':'sum', 'subtotal':'sum'}).reset_index()
            for idx, row in df_action.iterrows():
                q = int(row['qty'])
                amt = float(row['subtotal'])
                total_qty += q
                total_jumlah_semua += amt
                date_only = str(row['receipt_date'])[:10] 
                html_semua += f"<tr><td>{idx + 1}</td><td>{date_only}</td><td>{row['category_name']}</td><td>{row['action_name']}</td><td style='text-align: center;'>{q}</td><td style='text-align: right;'>{format_rupiah(amt).replace('Rp ', '')}</td></tr>"
                data_ex_semua.append({"No": idx + 1, "Tanggal Kuitansi": date_only, "Unit Layanan": row['category_name'], "Uraian Jenis Tindakan": row['action_name'], "Item": q, "Pemasukan (Rp)": amt})
        else:
            html_semua += "<tr><td colspan='6' style='text-align:center;'>Tidak ada data pada filter ini</td></tr>"

        html_semua += f"<tr class='total-row'><td colspan='4'>Total Pendapatan</td><td style='text-align: center;'>{total_qty}</td><td style='text-align: right;'>{format_rupiah(total_jumlah_semua).replace('Rp ', '')}</td></tr>"
        html_semua += "</table>"
        st.markdown(html_semua, unsafe_allow_html=True)
        df_ex_semua = pd.DataFrame(data_ex_semua)

        # =====================================================================
        # TAMPILAN: Rincian Penerimaan Piutang
        # =====================================================================
        st.markdown(f"#### Rincian Penerimaan Piutang ({str_date})")
        html_piu = "<table class='tbl-report'>"
        html_piu += "<tr><th style='width: 5%;'>No</th><th style='width: 25%;'>Nama Debitur / Pasien</th><th style='width: 25%;'>Metode Bayar</th><th style='width: 20%;'>Petugas</th><th style='width: 25%; text-align: right;'>Jumlah (Rp)</th></tr>"
        
        tot_piutang_masuk = 0.0
        if not df_piu.empty:
            for idx, row in df_piu.iterrows():
                p_amt = float(row['amount'] or 0)
                tot_piutang_masuk += p_amt
                d_name = row.get('debtor_name') or 'Pasien / Debitur'
                html_piu += f"<tr><td>{idx + 1}</td><td>{d_name}</td><td>{row.get('method', '-')}</td><td>{str(row.get('input_by', '-')).upper()}</td><td style='text-align: right;'>{format_rupiah(p_amt).replace('Rp ', '')}</td></tr>"
        else:
            html_piu += "<tr><td colspan='5' style='text-align:center;'>Tidak ada penerimaan piutang pada periode ini</td></tr>"
            
        html_piu += f"<tr class='total-row'><td colspan='4'>Total Penerimaan Piutang</td><td style='text-align: right;'>{format_rupiah(tot_piutang_masuk).replace('Rp ', '')}</td></tr>"
        html_piu += "</table>"
        st.markdown(html_piu, unsafe_allow_html=True)

        # =====================================================================
        # TAMPILAN: Summary Breakdown Dikelompokkan (Transaksi, Piutang, Uang Muka)
        # =====================================================================
        setoran = (tx_tunai + piu_tunai) + depo_tunai - pengembalian_depo_tunai

        html_summary = "<table class='summary-table'>"
        
        # Kelompok 1: Transaksi
        html_summary += "<tr><td colspan='2' style='background:#F1F5F9; font-weight:800; color:#0F172A;'>--- KELOMPOK TRANSAKSI ---</td></tr>"
        html_summary += f"<tr><td>Tunai (Transaksi)</td><td>{format_rupiah(tx_tunai).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>EDC (Transaksi)</td><td>{format_rupiah(tx_edc).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Qris (Transaksi)</td><td>{format_rupiah(tx_qris).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>VA (Transaksi)</td><td>{format_rupiah(tx_va).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Transfer (Transaksi)</td><td>{format_rupiah(tx_transfer).replace('Rp ', '')}</td></tr>"
        
        # Kelompok 2: Piutang
        html_summary += "<tr><td colspan='2' style='background:#EFF6FF; font-weight:800; color:#1E40AF;'>--- KELOMPOK PIUTANG ---</td></tr>"
        html_summary += f"<tr class='piutang-row'><td>Tunai (Pelunasan Piutang)</td><td>{format_rupiah(piu_tunai).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr class='piutang-row'><td>EDC (Pelunasan Piutang)</td><td>{format_rupiah(piu_edc).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr class='piutang-row'><td>Qris (Pelunasan Piutang)</td><td>{format_rupiah(piu_qris).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr class='piutang-row'><td>VA (Pelunasan Piutang)</td><td>{format_rupiah(piu_va).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr class='piutang-row'><td>Transfer (Pelunasan Piutang)</td><td>{format_rupiah(piu_transfer).replace('Rp ', '')}</td></tr>"
        
        # Kelompok 3: Uang Muka (Deposit)
        html_summary += "<tr><td colspan='2' style='background:#F8FAFC; font-weight:800; color:#0F172A;'>--- KELOMPOK UANG MUKA / DEPOSIT ---</td></tr>"
        html_summary += f"<tr><td>Uang Muka Tunai (Masuk)</td><td>{format_rupiah(depo_tunai).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Uang Muka EDC</td><td>{format_rupiah(depo_edc).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Uang Muka Transfer</td><td>{format_rupiah(depo_transfer).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Pengakuan / Penggunaan Uang Muka</td><td>{format_rupiah(tot_pengakuan).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td style='color:#EF4444;'>(-) Pengembalian Uang Muka Tunai</td><td>{format_rupiah(pengembalian_depo_tunai).replace('Rp ', '')}</td></tr>"
        html_summary += f"<tr><td>Pengembalian Uang Muka Transfer</td><td>{format_rupiah(pengembalian_depo_transfer).replace('Rp ', '')}</td></tr>"
        
        # Baris Setoran Real
        html_summary += f"<tr class='setoran-row'><td>Setoran Kasir Real ({current_kasir})</td><td>{format_rupiah(setoran).replace('Rp ', '')}</td></tr>"
        html_summary += "</table>"
        
        st.markdown(html_summary, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # =====================================================================
        # GENERATE FILE EXCEL DENGAN GARIS PEMBATAS (GRIDLINES)
        # =====================================================================
        output_excel = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Kasir"

        title_font = Font(size=14, bold=True)
        subtitle_font = Font(size=11, bold=True)
        header_fill = PatternFill(start_color="48BB78", end_color="48BB78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color="CBD5E1"), 
            right=Side(style='thin', color="CBD5E1"), 
            top=Side(style='thin', color="CBD5E1"), 
            bottom=Side(style='thin', color="CBD5E1")
        )

        def style_cell(cell, font=None, fill=None, align=None, border=None, num_format=None):
            if font: cell.font = font
            if fill: cell.fill = fill
            if align: cell.alignment = align
            if border: cell.border = border
            if num_format: cell.number_format = num_format

        ws['A1'] = 'RS SOEHARTO HEERDJAN'
        ws['A1'].font = title_font
        ws['A2'] = 'LAPORAN HARIAN PENERIMAAN KASIR'
        ws['A2'].font = subtitle_font
        ws['A3'] = f'Tanggal: {str_date}  |  Shift: {current_shift}  |  Kasir/User: {current_kasir}'

        curr_row = 5

        ws.cell(row=curr_row, column=1, value='Uraian Pemasukan Per Unit Layanan').font = subtitle_font
        curr_row += 1
        heads_unit = ['No', 'Kategori Layanan Utama', 'Unit Layanan', 'Penerimaan Jumlah (Rp)']
        for col_idx, h in enumerate(heads_unit, 1):
            c = ws.cell(row=curr_row, column=col_idx, value=h)
            style_cell(c, font=header_font, fill=header_fill, align=align_center, border=thin_border)
        curr_row += 1

        if not df_ex_unit.empty:
            for _, row in df_ex_unit.iterrows():
                ws.cell(row=curr_row, column=1, value=row['No']).border = thin_border
                ws.cell(row=curr_row, column=2, value=row['Kategori Layanan Utama']).border = thin_border
                ws.cell(row=curr_row, column=3, value=row['Unit Layanan']).border = thin_border
                style_cell(ws.cell(row=curr_row, column=4, value=row['Pemasukan (Rp)']), border=thin_border, num_format='#,##0')
                curr_row += 1

        ws.cell(row=curr_row, column=1, value='Total').font = Font(bold=True)
        ws.cell(row=curr_row, column=1).border = thin_border
        ws.cell(row=curr_row, column=2).border = thin_border
        ws.cell(row=curr_row, column=3).border = thin_border
        style_cell(ws.cell(row=curr_row, column=4, value=total_pendapatan), font=Font(bold=True), border=thin_border, num_format='#,##0')
        curr_row += 3

        ws.cell(row=curr_row, column=1, value=f'Uraian Semua Data (TRANSAKSI) - User: {display_user_meta} | Shift: {display_shift_meta}').font = subtitle_font
        curr_row += 1
        heads_tx = ['No', 'Tanggal Kuitansi', 'Unit Layanan', 'Uraian Jenis Tindakan', 'Item', 'Penerimaan Jumlah (Rp)']
        for col_idx, h in enumerate(heads_tx, 1):
            c = ws.cell(row=curr_row, column=col_idx, value=h)
            style_cell(c, font=header_font, fill=header_fill, align=align_center, border=thin_border)
        curr_row += 1

        if not df_ex_semua.empty:
            for _, row in df_ex_semua.iterrows():
                ws.cell(row=curr_row, column=1, value=row['No']).border = thin_border
                ws.cell(row=curr_row, column=2, value=row['Tanggal Kuitansi']).border = thin_border
                ws.cell(row=curr_row, column=3, value=row['Unit Layanan']).border = thin_border
                ws.cell(row=curr_row, column=4, value=row['Uraian Jenis Tindakan']).border = thin_border
                style_cell(ws.cell(row=curr_row, column=5, value=row['Item']), align=align_center, border=thin_border)
                style_cell(ws.cell(row=curr_row, column=6, value=row['Pemasukan (Rp)']), border=thin_border, num_format='#,##0')
                curr_row += 1

        ws.cell(row=curr_row, column=1, value='Total Pendapatan').font = Font(bold=True)
        ws.cell(row=curr_row, column=1).border = thin_border
        ws.cell(row=curr_row, column=2).border = thin_border
        ws.cell(row=curr_row, column=3).border = thin_border
        ws.cell(row=curr_row, column=4).border = thin_border
        style_cell(ws.cell(row=curr_row, column=5, value=total_qty), font=Font(bold=True), align=align_center, border=thin_border)
        style_cell(ws.cell(row=curr_row, column=6, value=total_jumlah_semua), font=Font(bold=True), border=thin_border, num_format='#,##0')
        curr_row += 3

        summary_excel_data = [
            ("Tunai (Transaksi)", tx_tunai),
            ("Tunai (Pelunasan Piutang)", piu_tunai),
            ("EDC (Transaksi)", tx_edc),
            ("EDC (Pelunasan Piutang)", piu_edc),
            ("Qris (Transaksi)", tx_qris),
            ("Qris (Pelunasan Piutang)", piu_qris),
            ("VA (Transaksi)", tx_va),
            ("VA (Pelunasan Piutang)", piu_va),
            ("Transfer (Transaksi)", tx_transfer),
            ("Transfer (Pelunasan Piutang)", piu_transfer),
            ("Uang Muka Tunai (Masuk)", depo_tunai),
            ("Uang Muka EDC", depo_edc),
            ("Uang Muka Transfer", depo_transfer),
            ("Pengakuan / Penggunaan Uang Muka", tot_pengakuan),
            ("Pengembalian Uang Muka Tunai", pengembalian_depo_tunai),
            ("Pengembalian Uang Muka Transfer", pengembalian_depo_transfer),
            (f"Setoran Kasir Real ({current_kasir})", setoran)
        ]
        
        for desc, val in summary_excel_data:
            is_bold = 'Setoran' in desc
            c_label = ws.cell(row=curr_row, column=5, value=desc)
            c_val = ws.cell(row=curr_row, column=6, value=val)
            
            style_cell(c_label, font=Font(bold=is_bold), align=align_right, border=thin_border)
            style_cell(c_val, font=Font(bold=is_bold), num_format='#,##0', border=thin_border)
            curr_row += 1

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25

        wb.save(output_excel)
        excel_data = output_excel.getvalue()

        # =====================================================================
        # TOMBOL AKSI CETAK / UNDUH
        # =====================================================================
        c_btn3, c_btn4, _ = st.columns([1.5, 1.5, 9])
        
        with c_btn3:
            st.markdown('<div class="btn-pdf">', unsafe_allow_html=True)
            if st.button("🖨️ Print / PDF", key="pdf_print", use_container_width=True):
                components.html("<script>window.parent.print();</script>", height=0)
            st.markdown('</div>', unsafe_allow_html=True)
            
        with c_btn4:
            st.markdown('<div class="btn-pdf">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Unduh Excel",
                data=excel_data,
                file_name=f"Laporan_Kasir_{current_kasir}_{str_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)